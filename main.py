import os
import subprocess
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logging.info("Script started successfully...")


INPUT_FOLDER = "inp"
OUTPUT_FOLDER = "output"
TEMPLATE_PATH = os.path.join(INPUT_FOLDER, "template.xlsx")
BALANCE_PATH = os.path.join(INPUT_FOLDER, "balance.xlsx")
TEMP_XLSX = os.path.join(OUTPUT_FOLDER, "temp.xlsx")
RAW_ODS = os.path.join(OUTPUT_FOLDER, "result_raw.ods")
FINAL_XLSX = os.path.join(OUTPUT_FOLDER, "result_raw.xlsx")

# Checks to perform
CHECKS = [
    {"sheet": "RC", "row": 38, "tol": 1.0, "msg": "SP attivo da AIDA"},
    {"sheet": "RC", "row": 89, "tol": 1.0, "msg": "SP passivo da AIDA"},
    {"sheet": "RC", "row": 90, "tol": 1.0, "msg": "SP attivo vs passivo"},
    {"sheet": "RC", "row": 134, "tol": 1.0, "msg": "reddito operativo da AIDA"},
    {"sheet": "RC", "row": 153, "tol": 1.0, "msg": "RN civ e ricl"},
    {"sheet": "RC", "row": 154, "tol": 1.0, "msg": "RN da CE vs SP"},
    {"sheet": "RF", "row": 56, "tol": 1.0, "msg": "sbilancio check NCF"},
    {"sheet": "RF", "row": 66, "tol": 1.0, "msg": "sbilancio check FCFF"},
    {"sheet": "RF", "row": 77, "tol": 1.0, "msg": "sbilancio check FCFE"},
]

def copy_balance_into_template():
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    balance_wb = load_workbook(BALANCE_PATH)
    template_wb = load_workbook(TEMPLATE_PATH)
    
    balance_ws = balance_wb.worksheets[0]
    template_ws = template_wb.worksheets[0]
    

    # Clear template sheet
    template_ws.delete_rows(1, template_ws.max_row)

    # Copy rows from balance sheet
    for row in balance_ws.iter_rows(values_only=True):
        template_ws.append(row)

    # Save as temp.xlsx then rename to .ods
    template_wb.save(TEMP_XLSX)
    os.rename(TEMP_XLSX, RAW_ODS)
    logging.info("Balance sheet copied to template and saved as %s", RAW_ODS)

def recalculate_with_libreoffice(input_path, output_dir):
    try:
        subprocess.run([
            "libreoffice",
            "--headless",
            "--convert-to", "xlsx",
            input_path,
            "--outdir", output_dir
        ], check=True)
        logging.info("LibreOffice recalculation complete.")
    except subprocess.CalledProcessError as e:
        logging.error("LibreOffice failed: %s", e)

def checkerror(wb, sheet_name, row_number, start_col=2, tolerance=1.0, friendly_message=None, treat_empty_as_zero=True):
    mismatches = []
    columns = []

    if sheet_name not in wb.sheetnames:
        logging.error("Sheet not found: %s", sheet_name)
        return {"mismatches": mismatches, "columns": columns, "summary": f"Sheet not found: {sheet_name}"}

    ws = wb[sheet_name]
    last_col = ws.max_column

    for col_idx in range(start_col, last_col + 1):
        cell = ws.cell(row=row_number, column=col_idx)
        raw = cell.value

        if raw is None:
            if treat_empty_as_zero:
                val = 0.0
            else:
                continue
        else:
            try:
                val = float(raw)
            except Exception:
                logging.error("Non-numeric value at %s row %d col %s: %r", sheet_name, row_number, get_column_letter(col_idx), raw)
                val = None

        numeric_val = 0.0 if (val is None and treat_empty_as_zero) else val

        if numeric_val is None:
            col_letter = get_column_letter(col_idx)
            mismatch = {"sheet": sheet_name, "row": row_number, "col": col_idx, "col_letter": col_letter, "raw": raw, "value": None}
            mismatches.append(mismatch)
            columns.append(col_letter)
            continue

        if abs(numeric_val) > tolerance:
            col_letter = get_column_letter(col_idx)
            mismatch = {"sheet": sheet_name, "row": row_number, "col": col_idx, "col_letter": col_letter, "raw": raw, "value": numeric_val}
            mismatches.append(mismatch)
            columns.append(col_letter)

    if columns:
        cols_str = ", ".join(columns)
        msg = friendly_message or f"Mismatch at {sheet_name} row {row_number}"
        summary = f"{msg} in columns: {cols_str}"
        logging.warning(summary)
    else:
        summary = None
        #logging.info("No mismatches for %s row %d", sheet_name, row_number)

    return {"mismatches": mismatches, "columns": columns, "summary": summary}

def run_checks(path):
    if not os.path.exists(path):
        logging.error("Recalculated file not found: %s", path)
        return

    wb = load_workbook(path, data_only=True)
    for chk in CHECKS:
        result = checkerror(
            wb,
            sheet_name=chk["sheet"],
            row_number=chk["row"],
            start_col=2,
            tolerance=chk["tol"],
            friendly_message=chk["msg"]
        )
        if result["summary"]:
            print(result["summary"])
        else:
            print(f"No mismatches for {chk['msg']}")

def main():
    copy_balance_into_template()
    recalculate_with_libreoffice(RAW_ODS, OUTPUT_FOLDER)
    run_checks(FINAL_XLSX)

if __name__ == "__main__":
    main()
